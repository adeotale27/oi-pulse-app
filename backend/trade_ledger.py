"""Persistent F&O trade cycles for the journal / Excel download.

Kite ``positions()`` has no fill clock. ``trades()`` is usually *today only*.
We therefore store each open cycle in Mongo the first time we see it, and
never invent a new entry time when the access token dies over the weekend.

Friday short held overnight → Monday reconnect with a fresh token is the
*same* cycle: purchase time stays Friday; exit is when the book actually
flattens. Partial closes are extra events on that cycle, not new trades.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

from kite_charges import parse_kite_timestamp
from market_hours import IST, now_ist
from universe import match_symbol_prefix

COLLECTION = "trade_cycles"


def instrument_key(row: Optional[dict]) -> Tuple[str, str, str]:
    if not isinstance(row, dict):
        return ("", "", "")
    return (
        str(row.get("exchange") or "").upper(),
        str(row.get("tradingsymbol") or "").upper(),
        str(row.get("product") or "").upper(),
    )


def iso_utc(dt: Optional[datetime] = None) -> str:
    d = dt or datetime.now(timezone.utc)
    if d.tzinfo is None:
        d = d.replace(tzinfo=timezone.utc)
    return d.astimezone(timezone.utc).isoformat()


def to_ist(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=IST)
    return dt.astimezone(IST)


def fmt_ist(dt: Optional[datetime]) -> Optional[str]:
    d = to_ist(dt)
    if d is None:
        return None
    return d.strftime("%Y-%m-%d %H:%M:%S")


def ymd_ist(dt: Optional[datetime] = None) -> str:
    d = to_ist(dt) or now_ist()
    return d.strftime("%Y-%m-%d")


def parse_dt(raw: Any) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return to_ist(raw)
    dt = parse_kite_timestamp(raw)
    return to_ist(dt) if dt else None


def _num(v: Any) -> float:
    try:
        n = float(v)
        return n if n == n else 0.0
    except (TypeError, ValueError):
        return 0.0


def _int(v: Any) -> int:
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def normalize_fill(raw: Optional[dict]) -> Optional[dict]:
    """Normalise a Kite trade (or COMPLETE order used as a fill)."""
    if not isinstance(raw, dict):
        return None
    ts = str(raw.get("tradingsymbol") or "").strip()
    if not ts:
        return None
    qty = _int(raw.get("quantity") or raw.get("filled_quantity"))
    if qty <= 0:
        return None
    ttype = str(raw.get("transaction_type") or raw.get("transactionType") or "").upper()
    if ttype not in ("BUY", "SELL"):
        return None
    when = parse_dt(
        raw.get("fill_timestamp")
        or raw.get("exchange_timestamp")
        or raw.get("order_timestamp")
        or raw.get("trade_timestamp")
    )
    if when is None:
        return None
    price = _num(raw.get("average_price") or raw.get("price"))
    tid = str(raw.get("trade_id") or raw.get("fill_id") or raw.get("order_id") or "").strip()
    return {
        "trade_id": tid or f"{ts}:{when.isoformat()}:{ttype}:{qty}:{price}",
        "exchange": str(raw.get("exchange") or "").upper(),
        "tradingsymbol": ts.upper(),
        "product": str(raw.get("product") or "").upper(),
        "transaction_type": ttype,
        "quantity": qty,
        "price": price,
        "time": when,
        "time_ist": fmt_ist(when),
        "time_iso": iso_utc(when),
    }


def collect_fills(trades: Optional[Iterable[dict]] = None, orders: Optional[Iterable[dict]] = None) -> List[dict]:
    out: List[dict] = []
    seen = set()
    rows = list(trades or [])
    if not rows:
        for o in orders or []:
            if str(o.get("status") or "").upper() in ("", "COMPLETE"):
                rows.append(o)
    for raw in rows:
        fill = normalize_fill(raw)
        if not fill:
            continue
        if fill["trade_id"] in seen:
            continue
        seen.add(fill["trade_id"])
        out.append(fill)
    out.sort(key=lambda f: f["time"] or datetime.min.replace(tzinfo=IST))
    return out


def fills_for_key(fills: Iterable[dict], key: Tuple[str, str, str]) -> List[dict]:
    ex, ts, prod = key
    matched = []
    for f in fills or []:
        if str(f.get("tradingsymbol") or "").upper() != ts:
            continue
        fex = str(f.get("exchange") or "").upper()
        if ex and fex and fex != ex:
            continue
        fprod = str(f.get("product") or "").upper()
        if prod and fprod and fprod != prod:
            continue
        matched.append(f)
    return matched


def new_cycle_id() -> str:
    return uuid.uuid4().hex


def _index_of(row: dict) -> str:
    raw = str(row.get("index") or "").strip().upper()
    if raw and raw not in ("OTHER", "UNKNOWN"):
        return raw
    ts = str(row.get("tradingsymbol") or row.get("display_name") or "")
    return match_symbol_prefix(ts) or raw or "OTHER"


def _direction(row: dict) -> str:
    qty = _int(row.get("quantity"))
    if qty < 0:
        return "short"
    if qty > 0:
        return "long"
    bias = str(row.get("side_bias") or "")
    if bias in ("short", "long", "squared"):
        return bias
    buy_q, sell_q = _int(row.get("buy_quantity")), _int(row.get("sell_quantity"))
    if sell_q > buy_q:
        return "short"
    if buy_q > sell_q:
        return "long"
    return "squared"


def _status(row: dict) -> str:
    if row.get("exited") or _int(row.get("quantity")) == 0:
        return "closed"
    if row.get("partial") or _int(row.get("closed_quantity")) > 0:
        return "partial"
    return "open"


def _opening_fills(fills: List[dict], direction: str) -> List[dict]:
    want = "SELL" if direction == "short" else "BUY"
    return [f for f in fills if f.get("transaction_type") == want]


def _closing_fills(fills: List[dict], direction: str) -> List[dict]:
    want = "BUY" if direction == "short" else "SELL"
    return [f for f in fills if f.get("transaction_type") == want]


def _event_ids(cycle: dict) -> set:
    ids = set()
    for ev in cycle.get("events") or []:
        tid = str(ev.get("trade_id") or "")
        if tid:
            ids.add(tid)
    return ids


def _append_event(
    cycle: dict,
    *,
    kind: str,
    fill: Optional[dict] = None,
    qty: int = 0,
    price: float = 0.0,
    remaining: int = 0,
    realised: float = 0.0,
    realised_this: Optional[float] = None,
    when: Optional[datetime] = None,
    note: str = "",
) -> None:
    events = list(cycle.get("events") or [])
    tid = (fill or {}).get("trade_id") or ""
    if tid and any(e.get("trade_id") == tid and e.get("kind") == kind for e in events):
        return
    t = (fill or {}).get("time") or when
    exited_qty = int(qty or (fill or {}).get("quantity") or 0)
    events.append({
        "kind": kind,
        "trade_id": tid,
        "quantity": exited_qty,
        "exited_quantity": exited_qty,
        "price": round(_num(price if price else (fill or {}).get("price")), 4),
        "time": iso_utc(t) if t else None,
        "time_ist": fmt_ist(t) if t else None,
        "remaining_quantity": int(remaining),
        "realised": round(_num(realised), 2),
        "realised_this": None if realised_this is None else round(_num(realised_this), 2),
        "note": note or "",
    })
    cycle["events"] = events
    _sync_partials(cycle)


def _sync_partials(cycle: dict) -> None:
    """First-class list of each scale-out: time, qty exited this fill, remaining, P&L slice."""
    partials = []
    running = 0.0
    seq = 0
    for ev in cycle.get("events") or []:
        if ev.get("kind") != "partial_exit":
            continue
        seq += 1
        slice_pnl = ev.get("realised_this")
        if slice_pnl is None:
            total = _num(ev.get("realised"))
            slice_pnl = round(total - running, 2)
            running = total
        else:
            running = round(running + _num(slice_pnl), 2)
        partials.append({
            "seq": seq,
            "time": ev.get("time"),
            "time_ist": ev.get("time_ist"),
            "exited_quantity": int(ev.get("exited_quantity") if ev.get("exited_quantity") is not None else ev.get("quantity") or 0),
            "remaining_quantity": int(ev.get("remaining_quantity") or 0),
            "price": ev.get("price"),
            "realised_this": round(_num(slice_pnl), 2),
            "realised_total": round(_num(ev.get("realised") if ev.get("realised") is not None else running), 2),
            "trade_id": ev.get("trade_id") or "",
            "note": ev.get("note") or "",
        })
    cycle["partials"] = partials
    cycle["partial_exit_count"] = len(partials)
    last = partials[-1] if partials else None
    cycle["last_partial_time_ist"] = (last or {}).get("time_ist")
    cycle["last_partial_qty"] = (last or {}).get("exited_quantity")
    cycle["partial_exited_quantity"] = sum(p["exited_quantity"] for p in partials)


def _stamp_times(cycle: dict) -> None:
    entry = parse_dt(cycle.get("entry_time"))
    exit_ = parse_dt(cycle.get("exit_time"))
    cycle["entry_time"] = iso_utc(entry) if entry else None
    cycle["entry_time_ist"] = fmt_ist(entry)
    cycle["entry_date"] = ymd_ist(entry) if entry else None
    cycle["exit_time"] = iso_utc(exit_) if exit_ else None
    cycle["exit_time_ist"] = fmt_ist(exit_)
    cycle["exit_date"] = ymd_ist(exit_) if exit_ else None


def seed_cycle(row: dict, *, owner_id: str, now: datetime, fills: Optional[List[dict]] = None) -> dict:
    key = instrument_key(row)
    direction = _direction(row)
    fills = fills or []
    overnight = _int(row.get("overnight_quantity"))
    opened = _opening_fills(fills, direction)
    closed = _closing_fills(fills, direction)
    first_open = opened[0] if opened else None
    last_close = closed[-1] if closed else None
    carried = abs(overnight) > 0
    if carried and not first_open:
        entry = now  # first observation; caller should prefer an existing cycle
        entry_source = "first_seen"
    elif first_open:
        entry = first_open["time"]
        entry_source = "fill"
    else:
        entry = now
        entry_source = "first_seen"
    status = _status(row)
    exit_dt = last_close["time"] if (status == "closed" and last_close) else (now if status == "closed" else None)
    exit_source = None
    if status == "closed":
        exit_source = "fill" if last_close else "flatten_seen"
    qty = 0 if status == "closed" else _int(row.get("quantity"))
    cycle = {
        "cycle_id": new_cycle_id(),
        "owner_id": owner_id,
        "exchange": key[0],
        "tradingsymbol": str(row.get("tradingsymbol") or ""),
        "display_name": row.get("display_name") or row.get("tradingsymbol"),
        "product": key[2],
        "index": _index_of(row),
        "side": row.get("side"),
        "strike": row.get("strike"),
        "direction": direction,
        "status": status,
        "carried": carried,
        "carried_from_date": ymd_ist(entry) if carried else None,
        "token_gap": False,
        "feed_stale": False,
        "quantity": qty,
        "overnight_quantity": overnight,
        "closed_quantity": _int(row.get("closed_quantity")),
        "buy_quantity": _int(row.get("buy_quantity")),
        "sell_quantity": _int(row.get("sell_quantity")),
        "average_price": round(_num(row.get("average_price")), 4),
        "entry_price": round(_num((first_open or {}).get("price") or row.get("average_price") or row.get("average_price_raw")), 4),
        "exit_price": round(_num((last_close or {}).get("price") or row.get("sell_price") or row.get("buy_price") or row.get("last_price")), 4) if status == "closed" else None,
        "last_price": round(_num(row.get("last_price")), 4),
        "realised": round(_num(row.get("realised") if row.get("realised") is not None else row.get("booked_pnl")), 2),
        "unrealised": round(_num(row.get("unrealised")), 2),
        "booked_pnl": round(_num(row.get("booked_pnl") if row.get("booked_pnl") is not None else row.get("realised")), 2),
        "entry_time": entry,
        "entry_source": entry_source,
        "exit_time": exit_dt,
        "exit_source": exit_source,
        "first_seen_at": iso_utc(now),
        "last_seen_at": iso_utc(now),
        "events": [],
        "updated_at": iso_utc(now),
    }
    if first_open:
        _append_event(cycle, kind="entry", fill=first_open, remaining=qty, when=first_open["time"])
    else:
        _append_event(cycle, kind="first_seen", qty=qty, price=cycle["entry_price"], remaining=qty, when=entry,
                      note="carried overnight" if carried else "")
    for f in opened[1:]:
        _append_event(cycle, kind="scale_in", fill=f, remaining=qty)
    prev_closed = 0
    for f in closed:
        prev_closed += _int(f.get("quantity"))
        kind = "exit" if status == "closed" and f is last_close else "partial_exit"
        _append_event(cycle, kind=kind, fill=f, remaining=qty if kind != "exit" else 0,
                      realised=cycle["booked_pnl"])
    _stamp_times(cycle)
    if carried:
        cycle["carried_from_date"] = cycle.get("entry_date")
    return cycle


def apply_row_to_cycle(
    cycle: dict,
    row: dict,
    *,
    now: datetime,
    fills: Optional[List[dict]] = None,
    today: Optional[str] = None,
) -> dict:
    """Update an existing cycle. Never replace a known entry_time."""
    out = dict(cycle)
    fills = fills or []
    today = today or ymd_ist(now)
    was_stale = bool(out.get("feed_stale"))
    direction = out.get("direction") or _direction(row)
    status = _status(row)
    qty = 0 if status == "closed" else _int(row.get("quantity"))
    overnight = _int(row.get("overnight_quantity"))
    closed_qty = _int(row.get("closed_quantity"))
    prev_closed = _int(out.get("closed_quantity"))
    prev_qty = _int(out.get("quantity"))
    prev_realised = _num(out.get("booked_pnl") if out.get("booked_pnl") is not None else out.get("realised"))

    if was_stale:
        out["token_gap"] = True
    out["feed_stale"] = False
    out["quantity"] = qty
    out["overnight_quantity"] = overnight
    out["closed_quantity"] = closed_qty
    out["buy_quantity"] = _int(row.get("buy_quantity"))
    out["sell_quantity"] = _int(row.get("sell_quantity"))
    out["average_price"] = round(_num(row.get("average_price")), 4)
    out["last_price"] = round(_num(row.get("last_price")), 4)
    out["realised"] = round(_num(row.get("realised") if row.get("realised") is not None else row.get("booked_pnl")), 2)
    out["unrealised"] = round(_num(row.get("unrealised")), 2)
    out["booked_pnl"] = round(_num(row.get("booked_pnl") if row.get("booked_pnl") is not None else row.get("realised")), 2)
    out["display_name"] = row.get("display_name") or out.get("display_name")
    out["index"] = out.get("index") or _index_of(row)
    out["side"] = out.get("side") or row.get("side")
    out["strike"] = out.get("strike") if out.get("strike") is not None else row.get("strike")
    out["last_seen_at"] = iso_utc(now)
    out["updated_at"] = iso_utc(now)

    if abs(overnight) > 0 or (out.get("entry_date") and out["entry_date"] < today):
        out["carried"] = True
        if not out.get("carried_from_date"):
            out["carried_from_date"] = out.get("entry_date") or today

    known = _event_ids(out)
    opened = _opening_fills(fills, direction)
    closed = _closing_fills(fills, direction)

    # Never clobber Friday's purchase clock with Monday's first poll.
    if not parse_dt(out.get("entry_time")):
        first_open = opened[0] if opened and not out.get("carried") else None
        entry = (first_open or {}).get("time") or now
        out["entry_time"] = entry
        out["entry_source"] = "fill" if first_open else "first_seen"
    elif out.get("entry_source") == "first_seen" and opened and not out.get("carried") and not was_stale:
        # Same-session open: upgrade first_seen to the actual fill once Kite sends it.
        first_open = opened[0]
        if first_open and ymd_ist(first_open["time"]) == today:
            out["entry_time"] = first_open["time"]
            out["entry_source"] = "fill"

    for f in opened:
        if f["trade_id"] in known:
            continue
        kind = "scale_in" if parse_dt(out.get("entry_time")) else "entry"
        _append_event(out, kind=kind, fill=f, remaining=qty)

    new_closes = [f for f in closed if f["trade_id"] not in known]
    qty_dropped = closed_qty > prev_closed or (
        status != "closed" and abs(qty) < abs(prev_qty) and prev_qty != 0
    )
    if status == "closed":
        partial_fills = new_closes[:-1] if len(new_closes) > 1 else []
        exit_fill = new_closes[-1] if new_closes else None
    else:
        partial_fills = new_closes if qty_dropped else []
        exit_fill = None
        if qty_dropped and not partial_fills:
            synthetic_qty = max(0, closed_qty - prev_closed) or max(0, abs(prev_qty) - abs(qty))
            if synthetic_qty:
                slice_pnl = round(_num(out["booked_pnl"]) - prev_realised, 2)
                _append_event(
                    out, kind="partial_exit", qty=synthetic_qty,
                    price=_num(row.get("buy_price") if direction == "short" else row.get("sell_price")),
                    remaining=qty, realised=out["booked_pnl"], realised_this=slice_pnl, when=now,
                    note="qty change vs last stored book",
                )

    delta = round(_num(out["booked_pnl"]) - prev_realised, 2)
    fill_qty_total = sum(_int(f.get("quantity")) for f in partial_fills) or 0
    if status == "closed" and exit_fill:
        fill_qty_total += _int(exit_fill.get("quantity"))
    remaining_left = abs(prev_qty)
    realised_acc = 0.0
    n_money = len(partial_fills) + (1 if (status == "closed" and (exit_fill or qty_dropped)) else 0)

    def _slice_pnl(fill_qty, idx, last_idx):
        if n_money <= 0:
            return 0.0
        if last_idx is not None and idx == last_idx:
            return round(delta - realised_acc, 2)
        if fill_qty_total > 0:
            return round(delta * (fill_qty / fill_qty_total), 2)
        return round(delta / n_money, 2)

    last_partial_i = len(partial_fills) - 1
    for i, f in enumerate(partial_fills):
        fq = _int(f.get("quantity"))
        remaining_left = max(0, remaining_left - fq)
        slice_pnl = _slice_pnl(fq, i, last_partial_i if status != "closed" else None)
        realised_acc = round(realised_acc + slice_pnl, 2)
        _append_event(
            out, kind="partial_exit", fill=f, qty=fq,
            remaining=remaining_left if status != "closed" else remaining_left,
            realised=round(prev_realised + realised_acc, 2),
            realised_this=slice_pnl,
        )

    if status == "closed":
        out["status"] = "closed"
        out["quantity"] = 0
        last_i = (len(partial_fills) if exit_fill else max(0, n_money - 1))
        slice_pnl = _slice_pnl(_int((exit_fill or {}).get("quantity")), last_i, last_i)
        if exit_fill:
            out["exit_time"] = exit_fill["time"]
            out["exit_source"] = "fill"
            out["exit_price"] = round(_num(exit_fill.get("price")), 4)
            _append_event(
                out, kind="exit", fill=exit_fill, remaining=0,
                realised=out["booked_pnl"], realised_this=slice_pnl,
            )
        else:
            out["exit_time"] = now
            out["exit_source"] = "inferred_after_stale" if was_stale else "flatten_seen"
            out["exit_price"] = round(_num(row.get("last_price") or row.get("buy_price") or row.get("sell_price")), 4)
            _append_event(
                out, kind="exit", qty=abs(prev_qty) or closed_qty, price=out["exit_price"],
                remaining=0, realised=out["booked_pnl"], realised_this=slice_pnl, when=now,
                note="flattened while token was stale" if was_stale else "flat on Kite book",
            )
    else:
        out["status"] = "partial" if (closed_qty > 0 or out.get("status") == "partial" or out.get("partials")) else "open"

    _stamp_times(out)
    _sync_partials(out)
    return out


def mark_cycles_stale(cycles: Iterable[dict], *, now: Optional[datetime] = None) -> List[dict]:
    """Token/feed dead: freeze clocks. Do not close or reopen anything."""
    now = now or now_ist()
    out = []
    for c in cycles or []:
        if not isinstance(c, dict):
            continue
        if c.get("status") == "closed":
            continue
        d = dict(c)
        d["feed_stale"] = True
        d["feed_stale_at"] = iso_utc(now)
        d["updated_at"] = iso_utc(now)
        out.append(d)
    return out


def close_missing_cycle(
    cycle: dict,
    *,
    now: datetime,
    fills: Optional[List[dict]] = None,
) -> dict:
    """Open in our DB, gone from Kite net — flattened while we may have been blind."""
    out = dict(cycle)
    was_stale = bool(out.get("feed_stale"))
    fills = fills or []
    direction = out.get("direction") or "short"
    closed = _closing_fills(fills, direction)
    last_close = closed[-1] if closed else None
    out["status"] = "closed"
    out["quantity"] = 0
    out["feed_stale"] = False
    if was_stale:
        out["token_gap"] = True
    out["last_seen_at"] = iso_utc(now)
    out["updated_at"] = iso_utc(now)
    if last_close:
        out["exit_time"] = last_close["time"]
        out["exit_source"] = "fill"
        out["exit_price"] = round(_num(last_close.get("price")), 4)
        _append_event(out, kind="exit", fill=last_close, remaining=0, realised=_num(out.get("booked_pnl")))
    else:
        last = parse_dt(out.get("last_seen_at")) or now
        out["exit_time"] = last
        out["exit_source"] = "inferred_after_stale" if was_stale else "flatten_seen"
        _append_event(
            out, kind="exit", qty=abs(_int(out.get("quantity"))),
            price=_num(out.get("last_price")), remaining=0,
            realised=_num(out.get("booked_pnl")), when=last,
            note="missing from Kite book after token refresh" if was_stale else "missing from Kite book",
        )
    _stamp_times(out)
    return out


def reconcile_cycles(
    existing: Iterable[dict],
    positions: Iterable[dict],
    fills: Optional[Iterable[dict]] = None,
    *,
    owner_id: str,
    now: Optional[datetime] = None,
    feed_ok: bool = True,
) -> List[dict]:
    """Return the cycles that must be upserted.

    When ``feed_ok`` is False (stale token), only mark open cycles stale.
    """
    now = now or now_ist()
    existing = [dict(c) for c in (existing or []) if isinstance(c, dict)]
    if not feed_ok:
        return mark_cycles_stale(existing, now=now)

    fill_list = list(fills or [])
    open_by_key: Dict[Tuple[str, str, str], dict] = {}
    for c in existing:
        if c.get("status") == "closed":
            continue
        open_by_key[instrument_key(c)] = c

    seen_keys = set()
    upserts: List[dict] = []
    today = ymd_ist(now)

    for row in positions or []:
        if not isinstance(row, dict):
            continue
        key = instrument_key(row)
        if not key[1]:
            continue
        seen_keys.add(key)
        row_fills = fills_for_key(fill_list, key)
        cur = open_by_key.get(key)
        if cur:
            upserts.append(apply_row_to_cycle(cur, row, now=now, fills=row_fills, today=today))
        else:
            # Same-day flatten with no stored open cycle still gets a closed row.
            if _status(row) == "closed" and _int(row.get("buy_quantity")) == 0 and _int(row.get("sell_quantity")) == 0:
                continue
            upserts.append(seed_cycle(row, owner_id=owner_id, now=now, fills=row_fills))

    for key, cur in open_by_key.items():
        if key in seen_keys:
            continue
        row_fills = fills_for_key(fill_list, key)
        upserts.append(close_missing_cycle(cur, now=now, fills=row_fills))

    return upserts


def cycle_in_range(cycle: dict, start: str, end: str) -> bool:
    if not cycle:
        return False
    entry = str(cycle.get("entry_date") or "")[:10]
    exit_ = str(cycle.get("exit_date") or "")[:10]
    if entry and start <= entry <= end:
        return True
    if exit_ and start <= exit_ <= end:
        return True
    if entry and entry <= end and (not exit_ or exit_ >= start):
        return True
    for ev in cycle.get("events") or []:
        t = str(ev.get("time_ist") or ev.get("time") or "")[:10]
        if t and start <= t <= end:
            return True
    return False


def filter_cycles(
    cycles: Iterable[dict],
    *,
    start: str,
    end: str,
    index: Optional[str] = None,
    status: Optional[str] = None,
) -> List[dict]:
    want = str(index or "").strip().upper()
    if want in ("", "ALL", "ALL_INDICES", "*"):
        want = None
    want_status = str(status or "").strip().lower()
    if want_status in ("", "all"):
        want_status = None
    out = []
    for c in cycles or []:
        if not cycle_in_range(c, start, end):
            continue
        if want and str(c.get("index") or "").upper() != want:
            ts = str(c.get("tradingsymbol") or "")
            if match_symbol_prefix(ts) != want:
                continue
        if want_status and str(c.get("status") or "") != want_status:
            if want_status == "open" and c.get("status") == "partial":
                pass
            else:
                continue
        out.append(c)
    out.sort(key=lambda c: (c.get("entry_time") or "", c.get("tradingsymbol") or ""))
    return out


def stamp_journal_legs(legs: List[dict], cycles: Iterable[dict]) -> bool:
    """Copy entry/exit clocks and partials onto journal snapshot legs."""
    by_ts: Dict[str, dict] = {}
    by_key: Dict[Tuple[str, str, str], dict] = {}
    for c in cycles or []:
        by_key[instrument_key(c)] = c
        ts = str(c.get("tradingsymbol") or "")
        if ts:
            by_ts[ts.upper()] = c
    changed = False
    for leg in legs or []:
        if not isinstance(leg, dict):
            continue
        c = by_key.get(instrument_key(leg)) or by_ts.get(str(leg.get("tradingsymbol") or "").upper())
        if not c:
            continue
        pairs = (
            ("entry_time_ist", "entry_time"),
            ("exit_time_ist", "exit_time"),
            ("entry_source", "entry_source"),
            ("exit_source", "exit_source"),
            ("carried", "carried"),
            ("token_gap", "token_gap"),
            ("status", "cycle_status"),
            ("partials", "partials"),
            ("partial_exit_count", "partial_exit_count"),
            ("last_partial_time_ist", "last_partial_time"),
            ("last_partial_qty", "last_partial_qty"),
        )
        for src, dst in pairs:
            val = c.get(src)
            if leg.get(dst) != val and val is not None:
                leg[dst] = val
                changed = True
    return changed
    if not doc:
        return None
    out = {k: v for k, v in doc.items() if k != "_id"}
    return out


EXPORT_HEADERS = [
    "Instrument",
    "Display",
    "Index",
    "Product",
    "Side",
    "Strike",
    "Direction",
    "Status",
    "Carried overnight",
    "Carried from",
    "Token gap",
    "Qty open",
    "Qty closed",
    "Entry time (IST)",
    "Exit time (IST)",
    "Entry source",
    "Exit source",
    "Entry price",
    "Exit / last price",
    "Realised P&L",
    "Unrealised P&L",
    "Booked P&L",
    "First seen (UTC)",
    "Last seen (UTC)",
    "Partial exits",
    "Last partial time (IST)",
    "Last partial qty",
]


def cycle_export_row(c: dict) -> list:
    return [
        c.get("tradingsymbol") or "",
        c.get("display_name") or "",
        c.get("index") or "",
        c.get("product") or "",
        c.get("side") or "",
        c.get("strike") if c.get("strike") is not None else "",
        c.get("direction") or "",
        c.get("status") or "",
        "Y" if c.get("carried") else "N",
        c.get("carried_from_date") or "",
        "Y" if c.get("token_gap") else "N",
        c.get("quantity") if c.get("quantity") is not None else "",
        c.get("closed_quantity") if c.get("closed_quantity") is not None else "",
        c.get("entry_time_ist") or "",
        c.get("exit_time_ist") or "",
        c.get("entry_source") or "",
        c.get("exit_source") or "",
        c.get("entry_price") if c.get("entry_price") is not None else "",
        c.get("exit_price") if c.get("exit_price") is not None else (c.get("last_price") or ""),
        c.get("booked_pnl") if c.get("booked_pnl") is not None else c.get("realised"),
        c.get("unrealised") if c.get("unrealised") is not None else "",
        c.get("booked_pnl") if c.get("booked_pnl") is not None else "",
        c.get("first_seen_at") or "",
        c.get("last_seen_at") or "",
        c.get("partial_exit_count") or 0,
        c.get("last_partial_time_ist") or "",
        c.get("last_partial_qty") if c.get("last_partial_qty") is not None else "",
    ]


EVENT_HEADERS = [
    "Instrument",
    "Index",
    "Kind",
    "Time (IST)",
    "Exited qty this fill",
    "Remaining qty",
    "Price",
    "Realised this fill",
    "Realised total after fill",
    "Trade id",
    "Note",
]


def event_export_rows(cycles: Iterable[dict]) -> List[list]:
    rows = []
    for c in cycles or []:
        for ev in c.get("events") or []:
            rows.append([
                c.get("tradingsymbol") or "",
                c.get("index") or "",
                ev.get("kind") or "",
                ev.get("time_ist") or "",
                ev.get("exited_quantity") if ev.get("exited_quantity") is not None else ev.get("quantity"),
                ev.get("remaining_quantity") if ev.get("remaining_quantity") is not None else "",
                ev.get("price") if ev.get("price") is not None else "",
                ev.get("realised_this") if ev.get("realised_this") is not None else "",
                ev.get("realised") if ev.get("realised") is not None else "",
                ev.get("trade_id") or "",
                ev.get("note") or "",
            ])
    return rows


def partial_export_rows(cycles: Iterable[dict]) -> List[list]:
    rows = []
    for c in cycles or []:
        for p in c.get("partials") or []:
            rows.append([
                c.get("tradingsymbol") or "",
                c.get("index") or "",
                p.get("seq"),
                p.get("time_ist") or "",
                p.get("exited_quantity"),
                p.get("remaining_quantity"),
                p.get("price"),
                p.get("realised_this"),
                p.get("realised_total"),
            ])
    return rows


def workbook_bytes(cycles: List[dict], *, start: str, end: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="065F46")
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for c in cycles:
        ws.append(cycle_export_row(c))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, _h in enumerate(EXPORT_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["N"].width = 20
    ws.column_dimensions["O"].width = 20

    ev = wb.create_sheet("Fills and partials")
    ev.append(EVENT_HEADERS)
    for cell in ev[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in event_export_rows(cycles):
        ev.append(row)
    ev.freeze_panes = "A2"
    ev.auto_filter.ref = ev.dimensions
    for i, _h in enumerate(EVENT_HEADERS, 1):
        ev.column_dimensions[get_column_letter(i)].width = 16
    ev.column_dimensions["A"].width = 26
    ev.column_dimensions["D"].width = 20

    pr = wb.create_sheet("Partials")
    pr_headers = [
        "Instrument", "Index", "#", "Time (IST)", "Exited qty this fill",
        "Remaining qty", "Price", "Realised this fill", "Realised total",
    ]
    pr.append(pr_headers)
    for cell in pr[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in partial_export_rows(cycles):
        pr.append(row)
    pr.freeze_panes = "A2"
    pr.auto_filter.ref = pr.dimensions
    for i, _h in enumerate(pr_headers, 1):
        pr.column_dimensions[get_column_letter(i)].width = 18
    pr.column_dimensions["A"].width = 26
    pr.column_dimensions["D"].width = 20

    meta = wb.create_sheet("Notes")
    meta.append(["Filter from", start])
    meta.append(["Filter to", end])
    meta.append(["Rows", len(cycles)])
    meta.append(["Timezone", "Asia/Kolkata (IST) for entry/exit clocks"])
    meta.append([
        "Carried overnight",
        "Y means this cycle was still open after a session (e.g. Friday hold → Monday). Entry stays the original purchase time.",
    ])
    meta.append([
        "Token gap",
        "Y means Kite access token was stale while this cycle was open. We do not treat Monday reconnect as a new trade.",
    ])
    meta.append([
        "Partial exits",
        "Each scale-out is a row on the Partials sheet: time, how much was exited this fill, remaining qty, and realised for that slice. The parent trade keeps one entry time until fully flattened.",
    ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
