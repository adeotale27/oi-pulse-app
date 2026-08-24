"""Trade ledger: carry-forward clocks, stale tokens, partial exits, Excel."""
from datetime import datetime, timedelta, timezone

from trade_ledger import (
    apply_row_to_cycle,
    close_missing_cycle,
    collect_fills,
    cycle_in_range,
    filter_cycles,
    instrument_key,
    mark_cycles_stale,
    reconcile_cycles,
    seed_cycle,
    stamp_journal_legs,
    workbook_bytes,
    ymd_ist,
)

IST = timezone(timedelta(hours=5, minutes=30))


def _dt(y, m, d, hh, mm, ss=0):
    return datetime(y, m, d, hh, mm, ss, tzinfo=IST)


def _row(**kwargs):
    base = {
        "exchange": "NFO",
        "tradingsymbol": "NIFTY2582124500CE",
        "display_name": "NIFTY 24500 CE",
        "product": "NRML",
        "index": "NIFTY",
        "side": "CE",
        "strike": 24500,
        "quantity": -75,
        "overnight_quantity": 0,
        "average_price": 120.5,
        "last_price": 80.0,
        "pnl": 3000,
        "realised": 0,
        "unrealised": 3000,
        "booked_pnl": 0,
        "buy_quantity": 0,
        "sell_quantity": 75,
        "buy_price": 0,
        "sell_price": 120.5,
        "exited": False,
        "partial": False,
        "closed_quantity": 0,
        "side_bias": "short",
    }
    base.update(kwargs)
    return base


def _fill(**kwargs):
    base = {
        "trade_id": "t1",
        "exchange": "NFO",
        "tradingsymbol": "NIFTY2582124500CE",
        "product": "NRML",
        "transaction_type": "SELL",
        "quantity": 75,
        "average_price": 120.5,
        "fill_timestamp": "2026-08-21 14:32:11",
    }
    base.update(kwargs)
    return base


def test_seed_uses_fill_time_as_entry():
    friday = _dt(2026, 8, 21, 14, 40)
    fills = collect_fills([_fill()])
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=fills)
    assert c["entry_source"] == "fill"
    assert c["entry_time_ist"] == "2026-08-21 14:32:11"
    assert c["exit_time"] is None
    assert c["status"] == "open"
    assert c["direction"] == "short"


def test_partial_exit_keeps_entry_and_no_exit_clock():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    later = _dt(2026, 8, 21, 15, 10)
    partial = _row(
        quantity=-50,
        partial=True,
        closed_quantity=25,
        buy_quantity=25,
        sell_quantity=75,
        realised=800,
        booked_pnl=800,
        pnl=2200,
    )
    fills = collect_fills([
        _fill(),
        _fill(trade_id="t2", transaction_type="BUY", quantity=25, average_price=88.0,
              fill_timestamp="2026-08-21 15:09:01"),
    ])
    out = apply_row_to_cycle(c, partial, now=later, fills=fills, today="2026-08-21")
    assert out["entry_time_ist"] == "2026-08-21 14:32:11"
    assert out["exit_time"] is None
    assert out["status"] == "partial"
    kinds = [e["kind"] for e in out["events"]]
    assert "partial_exit" in kinds
    assert out["closed_quantity"] == 25
    pe = out["partials"]
    assert len(pe) == 1
    assert pe[0]["exited_quantity"] == 25
    assert pe[0]["remaining_quantity"] == 50
    assert pe[0]["time_ist"] == "2026-08-21 15:09:01"
    assert pe[0]["realised_this"] == 800
    assert out["last_partial_qty"] == 25
    assert out["last_partial_time_ist"] == "2026-08-21 15:09:01"


def test_two_partials_each_store_time_and_qty():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    first = apply_row_to_cycle(
        c,
        _row(quantity=-50, partial=True, closed_quantity=25, buy_quantity=25, sell_quantity=75,
             realised=800, booked_pnl=800),
        now=_dt(2026, 8, 21, 15, 10),
        fills=collect_fills([
            _fill(),
            _fill(trade_id="p1", transaction_type="BUY", quantity=25, average_price=88.0,
                  fill_timestamp="2026-08-21 15:09:01"),
        ]),
        today="2026-08-21",
    )
    second = apply_row_to_cycle(
        first,
        _row(quantity=-25, partial=True, closed_quantity=50, buy_quantity=50, sell_quantity=75,
             realised=1100, booked_pnl=1100),
        now=_dt(2026, 8, 21, 15, 25),
        fills=collect_fills([
            _fill(),
            _fill(trade_id="p1", transaction_type="BUY", quantity=25, average_price=88.0,
                  fill_timestamp="2026-08-21 15:09:01"),
            _fill(trade_id="p2", transaction_type="BUY", quantity=25, average_price=90.0,
                  fill_timestamp="2026-08-21 15:24:40"),
        ]),
        today="2026-08-21",
    )
    assert len(second["partials"]) == 2
    assert second["partials"][0]["exited_quantity"] == 25
    assert second["partials"][0]["time_ist"] == "2026-08-21 15:09:01"
    assert second["partials"][1]["exited_quantity"] == 25
    assert second["partials"][1]["remaining_quantity"] == 25
    assert second["partials"][1]["time_ist"] == "2026-08-21 15:24:40"
    assert second["partials"][1]["realised_this"] == 300
    assert second["entry_time_ist"] == "2026-08-21 14:32:11"
    assert second["exit_time"] is None


def test_full_exit_sets_exit_from_fill():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    done = _row(
        quantity=0, exited=True, closed_quantity=75,
        buy_quantity=75, sell_quantity=75, realised=2100, booked_pnl=2100, pnl=2100,
        average_price=0, side_bias="squared",
    )
    fills = collect_fills([
        _fill(),
        _fill(trade_id="t-exit", transaction_type="BUY", quantity=75, average_price=92.5,
              fill_timestamp="2026-08-21 15:20:44"),
    ])
    out = apply_row_to_cycle(c, done, now=_dt(2026, 8, 21, 15, 21), fills=fills, today="2026-08-21")
    assert out["status"] == "closed"
    assert out["exit_time_ist"] == "2026-08-21 15:20:44"
    assert out["exit_source"] == "fill"
    assert out["entry_time_ist"] == "2026-08-21 14:32:11"


def test_stale_token_does_not_close_or_reset_entry():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    monday_morn = _dt(2026, 8, 24, 7, 5)
    stale = reconcile_cycles([c], positions=[], fills=[], owner_id="admin", now=monday_morn, feed_ok=False)
    assert len(stale) == 1
    assert stale[0]["feed_stale"] is True
    assert stale[0]["status"] == "open"
    assert stale[0]["entry_time_ist"] == "2026-08-21 14:32:11"
    assert stale[0].get("exit_time") in (None, c.get("exit_time"))


def test_monday_reconnect_keeps_friday_entry_and_marks_carried():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    c = mark_cycles_stale([c], now=_dt(2026, 8, 24, 7, 0))[0]
    monday = _dt(2026, 8, 24, 9, 20)
    live = _row(overnight_quantity=-75, quantity=-75)
    out = reconcile_cycles(
        [c], [live], fills=[], owner_id="admin", now=monday, feed_ok=True,
    )[0]
    assert out["entry_time_ist"] == "2026-08-21 14:32:11"
    assert out["entry_source"] == "fill"
    assert out["carried"] is True
    assert out["token_gap"] is True
    assert out["feed_stale"] is False
    assert out["status"] == "open"
    assert out["exit_time"] is None


def test_monday_exit_of_friday_hold_uses_monday_fill_for_exit_only():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    c = mark_cycles_stale([c], now=_dt(2026, 8, 24, 7, 0))[0]
    monday = _dt(2026, 8, 24, 10, 5)
    flat = _row(
        quantity=0, exited=True, overnight_quantity=-75, closed_quantity=75,
        buy_quantity=75, sell_quantity=75, realised=1500, booked_pnl=1500, pnl=1500,
        average_price=0, side_bias="squared",
    )
    fills = collect_fills([
        _fill(trade_id="mon-buy", transaction_type="BUY", quantity=75, average_price=100.0,
              fill_timestamp="2026-08-24 10:04:12"),
    ])
    out = reconcile_cycles([c], [flat], fills, owner_id="admin", now=monday, feed_ok=True)[0]
    assert out["entry_time_ist"] == "2026-08-21 14:32:11"
    assert out["exit_time_ist"] == "2026-08-24 10:04:12"
    assert out["carried"] is True
    assert out["token_gap"] is True
    assert out["status"] == "closed"


def test_closed_during_stale_gap_inferred_from_missing_book():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    c["last_seen_at"] = "2026-08-21T10:10:00+00:00"
    c = mark_cycles_stale([c], now=_dt(2026, 8, 21, 15, 50))[0]
    monday = _dt(2026, 8, 24, 9, 20)
    out = reconcile_cycles([c], positions=[], fills=[], owner_id="admin", now=monday, feed_ok=True)[0]
    assert out["status"] == "closed"
    assert out["entry_time_ist"] == "2026-08-21 14:32:11"
    assert out["exit_source"] == "inferred_after_stale"
    assert out["token_gap"] is True


def test_new_cycle_after_previous_close_same_symbol():
    friday = _dt(2026, 8, 21, 14, 40)
    closed = seed_cycle(
        _row(quantity=0, exited=True, buy_quantity=75, sell_quantity=75, closed_quantity=75,
             realised=100, booked_pnl=100, average_price=0),
        owner_id="admin", now=friday,
        fills=collect_fills([
            _fill(),
            _fill(trade_id="x", transaction_type="BUY", quantity=75, fill_timestamp="2026-08-21 15:00:00"),
        ]),
    )
    assert closed["status"] == "closed"
    monday = _dt(2026, 8, 24, 11, 0)
    fresh = _row(quantity=-150, sell_quantity=150, overnight_quantity=0)
    fills = collect_fills([
        _fill(trade_id="new", quantity=150, fill_timestamp="2026-08-24 10:55:00"),
    ])
    out = reconcile_cycles([closed], [fresh], fills, owner_id="admin", now=monday, feed_ok=True)
    assert len(out) == 1
    assert out[0]["cycle_id"] != closed["cycle_id"]
    assert out[0]["entry_time_ist"] == "2026-08-24 10:55:00"
    assert out[0]["status"] == "open"


def test_filter_includes_carried_open_in_later_range():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    c["carried"] = True
    assert cycle_in_range(c, "2026-08-24", "2026-08-24") is True
    assert cycle_in_range(c, "2026-08-10", "2026-08-12") is False
    rows = filter_cycles([c], start="2026-08-21", end="2026-08-24", index="NIFTY")
    assert len(rows) == 1
    assert filter_cycles([c], start="2026-08-21", end="2026-08-24", index="SENSEX") == []


def test_stamp_journal_legs_copies_clocks():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    legs = [{"tradingsymbol": "NIFTY2582124500CE", "product": "NRML", "exchange": "NFO"}]
    assert stamp_journal_legs(legs, [c]) is True
    assert legs[0]["entry_time"] == "2026-08-21 14:32:11"
    assert legs[0]["carried"] is False


def test_workbook_has_entry_and_exit_columns():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    raw = workbook_bytes([c], start="2026-08-21", end="2026-08-24")
    assert raw[:2] == b"PK"
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(BytesIO(raw))
    ws = wb["Trades"]
    headers = [cell.value for cell in ws[1]]
    assert "Entry time (IST)" in headers
    assert "Exit time (IST)" in headers
    ei = headers.index("Entry time (IST)")
    assert ws[2][ei].value == "2026-08-21 14:32:11"
    assert "Fills and partials" in wb.sheetnames
    assert "Partials" in wb.sheetnames


def test_instrument_key_and_ymd():
    assert instrument_key(_row()) == ("NFO", "NIFTY2582124500CE", "NRML")
    assert ymd_ist(_dt(2026, 8, 21, 23, 10)) == "2026-08-21"


def test_close_missing_uses_today_fill_when_present():
    friday = _dt(2026, 8, 21, 14, 40)
    c = seed_cycle(_row(), owner_id="admin", now=friday, fills=collect_fills([_fill()]))
    c = mark_cycles_stale([c], now=_dt(2026, 8, 24, 7, 0))[0]
    fills = collect_fills([
        _fill(trade_id="buy", transaction_type="BUY", quantity=75, average_price=99,
              fill_timestamp="2026-08-24 09:16:00"),
    ])
    out = close_missing_cycle(c, now=_dt(2026, 8, 24, 9, 20), fills=fills)
    assert out["exit_time_ist"] == "2026-08-24 09:16:00"
    assert out["exit_source"] == "fill"


def test_summarize_trade_memory_weekday_and_min_n():
    from trade_ledger import summarize_trade_memory

    cycles = []
    for i in range(5):
        cycles.append({
            "status": "closed",
            "direction": "short",
            "index": "NIFTY",
            "side": "CE",
            "booked_pnl": 200 if i < 4 else -80,
            "entry_date": "2026-08-21",
        })
    cycles.append({
        "status": "closed",
        "direction": "short",
        "index": "NIFTY",
        "side": "PE",
        "booked_pnl": 10,
        "entry_date": "2026-08-21",
    })
    mem = summarize_trade_memory(cycles)
    assert any("NIFTY CE shorts on Friday: 4/5 paid" in x for x in mem["lines"])
    skinny = summarize_trade_memory(cycles[:2])
    assert skinny["lines"] == []
